import streamlit as st
import time
import requests
import base64
import pytz
from fake_useragent import UserAgent
import openpyxl
from bs4 import BeautifulSoup
import re
from datetime import datetime, timedelta
import os

def monday_run():
    
    session = requests.Session()

    #Define headers
    headers = {
        'User-Agent': UserAgent().windows,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en-US,en;q=0.9',
        'Connection': 'keep-alive',
        'DNT': '1',  # Do Not Track request header
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://amazon.in/',
    }

    session.cookies.set('session-id', '259-9131690-4965843', domain='.amazon.in')
    session.cookies.set('ubid-acbin', '262-3136477-5925466', domain='.amazon.in')
    session.cookies.set('x-acbin', 'oKxIUSyFo?JHcckDwUY?v@ZYnayTGSw8@bjN1GChjq0wUS@I8w31mOz?cHcIzdqI', domain='.amazon.in')

    sku_sheet = openpyxl.load_workbook("testSKU.xlsx").active
    marketplace_review_workbook = openpyxl.Workbook()
    default_sheet = marketplace_review_workbook.active
    marketplace_review_workbook.remove(default_sheet)
    marketplace_review = marketplace_review_workbook.create_sheet(title="Review")
    head_mp = ["ECOM SKU","SKU Name","ASIN", "Marketplace", "Name", "Title","Review", "Rating", "Date", "Location", "Sentiment","Images"]
    marketplace_negative = marketplace_review_workbook.create_sheet(title="Negative review in last 3 days")
    head_ng = ["ECOM SKU","SKU Name","ASIN", "Marketplace", "Name", "Title","Review", "Rating", "Date", "Location", "Images"]

    marketplace_review.append(head_mp)
    marketplace_negative.append(head_ng)

    mp_review = marketplace_review_workbook.create_sheet(title="Rating")
    heading = ['ECOM SKU ID','SKU Name','ASIN/FSN','Marketplace','Status','Rating 5','Rating 4','Rating 3','Rating 2','Rating 1','Average Rating','Total Ratings','Total Reviews','Last Review Date','Review count in last 30 days','Review count in last 365 days']
    filename = f"Amazon_review_{current_time}.xlsx"
    mp_review.append(heading)

    for row in range(2,sku_sheet.max_row+1):
        sku = sku_sheet.cell(row=row,column=1).value
        amz_mark = sku_sheet.cell(row=row,column=3).value
        col = marketplace_review.max_row+1
        if amz_mark == 'Y':
            asin = sku_sheet.cell(row=row,column=5).value
            print(f'/// ASIN {asin}')
            #Load asin page
            url = f'https://www.amazon.in/dp/{asin}'
            print(f'Load {url}...')
            response = session.get(url, headers=headers, verify=False)
            headers['Referer'] = url #save url as referrer for next query
            print(response)
            soup_mp = BeautifulSoup(response.content,"html.parser")
            sku_name = soup_mp.find('span', {'id': 'productTitle'}).get_text(strip=True)
            
            status_tag = soup_mp.find('span', {'class':'a-size-medium a-color-success'}) 
            if status_tag:
                status = status_tag.get_text(strip=True) 
            else:
                status = "Avaiable"
            print(sku_name)
            marketplace = "Amazon"
        
            # Find the div containing the rating histogram
            rating_histogram_div = soup_mp.find('div', class_='a-section a-spacing-none a-text-right aok-nowrap')

            # Initialize variables to store the ratings
            rating_5 = rating_4 = rating_3 = rating_2 = rating_1 = None

            if rating_histogram_div:
                # Extract the percentages from the span elements inside the div
                spans = rating_histogram_div.find_all('span', class_='_cr-ratings-histogram_style_histogram-column-space__RKUAd')
                
                # Store the ratings in corresponding variables
                if len(spans) >= 5:
                    rating_5 = spans[0].text.strip()  # First span: 51%
                    rating_4 = spans[1].text.strip()  # Second span: 18%
                    rating_3 = spans[2].text.strip()  # Third span: 3%
                    rating_2 = spans[3].text.strip()  # Fourth span: 2%
                    rating_1 = spans[4].text.strip()  # Fifth span: 26%
                
                # If there's additional text (like the final 51%) outside the spans, handle it
                last_rating = rating_histogram_div.contents[-1].strip()
                if last_rating and last_rating not in [rating_5, rating_4, rating_3, rating_2, rating_1]:
                    rating_5 = last_rating  # Assign it to rating_5 if not already set

            # Print the extracted ratings
            print(f'Rating 5: {rating_5}')
            print(f'Rating 4: {rating_4}')
            print(f'Rating 3: {rating_3}')
            print(f'Rating 2: {rating_2}')
            print(f'Rating 1: {rating_1}')
            avg_rating = soup_mp.find('span', {'data-hook': 'rating-out-of-text', 'aria-hidden': 'true'}).text.strip().split(' ')[0] if soup_mp.find('span', {'data-hook': 'rating-out-of-text', 'aria-hidden': 'true'}) else None
            print(avg_rating)
            number_of_rating = 0
            number_of_review = 0
            time.sleep(2)
            check_pr = soup_mp.find('span',{'data-hook' : 'total-review-count'})
            first_review_date = 101
            count_30 = 0
            count_365 = 0
            if check_pr:
                print("Running")
                page = 1
                while True:
                    url = f'https://www.amazon.in/product-reviews/{asin}/ref=cm_cr_arp_d_viewopt_srt?sortBy=recent&pageNumber={page}'
                    print(f'Page {page} | Load {url}...')
                    response_rw = session.get(url, headers=headers, verify=False)
                    headers['Referer'] = url  # save url as referrer for next query
                    time.sleep(2)
                    soup_rw = BeautifulSoup(response_rw.content,'html.parser')
                    reviews = soup_rw.findAll('li',{'class' : 'review aok-relative'})
                    # print(reviews)
                    if len(reviews) == 0:
                        break
                    
                    # no_review = soup_rw.find('div',{'class':'a-row a-spacing-base a-size-base'}).get_text(strip=True)

                    text = soup_rw.find('div', {'data-hook': 'cr-filter-info-review-rating-count'}).text.strip()
                    # Use regex to extract numbers
                    match = re.search(r'([\d,]+)\s+total rating(?:s)?,?\s+([\d,]+)?\s+with review(?:s)?', text)

                    number_of_rating = int(match.group(1).replace(',', ''))
                    number_of_review = int(match.group(2).replace(',', '')) if match.group(2) else 0
                    print(f"Number of ratings: {number_of_rating}")
                    print(f"Number of reviews: {number_of_review}")

                    # Use regex to extract numbers
                    # number_of_rating, number_of_review = text.split(' total ratings, ') if text.split(' total ratings, ') else text.split(' total rating, ')

                    # number_of_review = number_of_review.split(' with reviews')[0] or number_of_review.split(' with review')[0]

                    # Use string splitting to get the rating and review numbers
                    # number_of_rating, number_of_review = no_review.split(' total ratings, ')

                    # # Convert them to integers if needed
                    # number_of_rating = int(number_of_rating.replace(",", ""))
                    # number_of_review = int(number_of_review.split(' with reviews')[0]) 
                    print(number_of_rating)
                    print(number_of_review)
                    for review in reviews:

                        name = review.find('span',{'class':'a-profile-name'}).text.strip()
                        rating = int(float(review.find('span', class_='a-icon-alt').text.split(' ')[0]))
                        title = review.find('a', {'data-hook': 'review-title'}).find_all('span')[-1].text.strip()
                        review_text = review.find('span',{'data-hook':'review-body'}).find('span').text.strip() if review.find('span',{'data-hook':'review-body'}).find('span') else "NA"
                        review_date_text = review.find('span', {'data-hook': 'review-date'}).text.strip()

                        # Use regex to extract location and date
                        location_match = re.search(r'Reviewed in (\w+)', review_date_text)
                        date_match = re.search(r'on (\d{1,2} \w+ \d{4})', review_date_text)

                        # Assign to variables if matched
                        location = location_match.group(1) if location_match else None
                        date = date_match.group(1) if date_match else None
                        if first_review_date == 101:
                            first_review_date = date
                        date_obj = datetime.strptime(date, "%d %B %Y")
                        today = datetime.today()

                        # Calculate date ranges
                        thirty_days_ago = today - timedelta(days=30)
                        three_hundred_sixty_five_days_ago = today - timedelta(days=365)

                        # Check if the date falls within the last 30 days or 365 days
                        if thirty_days_ago <= date_obj <= today:
                            count_30 += 1
                        if three_hundred_sixty_five_days_ago <= date_obj <= today:
                            count_365 += 1
                        print(name+' '+str(rating)+' '+date+' '+location)
                        image_tags = review.find_all('img', class_='review-image-tile') if review.find_all('img', class_='review-image-tile') else None
                        all_links = "NA"
                        if image_tags is not None:
                            # Extract the 'src' links
                            image_links = [img['src'].replace('SY88', 'SL1600') for img in image_tags]

                            # Join the links into a single string with a delimiter (comma in this case)
                            all_links = ', '.join(image_links)

                        # Now you have all links in one variable
                        print(all_links)
                        sentiment=""
                        if rating <= 3:
                            sentiment = "Negative"
                        else:
                            sentiment = "Positive"
                        marketplace_review.append([
                            sku,
                            sku_name,
                            asin,
                            marketplace,
                            name,
                            title,
                            review_text,
                            rating,
                            date,
                            location,
                            sentiment,
                            all_links
                        ])
                        if rating <= 3 and today - timedelta(days=3) <= date_obj:
                            marketplace_negative.append([
                                sku,
                                sku_name,
                                asin,
                                marketplace,
                                name,
                                title,
                                review_text,
                                rating,
                                date,
                                location,
                                all_links
                            ])

                    next_page_check = soup_rw.find('li',{'class':'a-last'})
                    if next_page_check:
                        next_page = next_page_check.find('a')
                        if next_page:
                            page +=1
                        else:
                            break
                    else:
                        break
            mp_review.append([
                sku,
                sku_name,
                asin,
                marketplace,
                status,
                rating_5,
                rating_4,
                rating_3,
                rating_2,
                rating_1,
                avg_rating,
                number_of_rating,
                number_of_review,
                first_review_date,
                count_30,
                count_365
            ])     
            marketplace_review_workbook.save(filename)

    with open(marketplace_review_workbook, mode='rb') as file:
        excel_base64 = base64.b64encode(file.read()).decode('utf-8')

    payload = {
        "personalizations": [
            {
                "recipient": "sparmar@godrej.com",
                "recipient_cc": ["yashkhot@godrej.com", "aarushid@godrej.com"]
            }
        ],
        "from": {
            "fromEmail": "info@godrejinterio.com",
        },
        "subject": "Amazon Ratings and Reviews",
        "content": "Hi, PFA the Ratings and Reviews Dump for Amazon",
        "templateId": 33201,
        "attachments": [
            {
                "fileContent": excel_base64,
                "fileName": "RatingsAndReviews.xlsx"
            }
        ]
    }

    headers = {
        "api_key": "50b39b265e9b8cca8a264652a27b57ef",
        "Content-Type": "application/json",
        "Accept": ""
    }
    url_email = "https://api.pepipost.com/v2/sendEmail"
    response_email = requests.post(url_email, json=payload, headers=headers)
    if os.path.exists(filename):
        os.remove(filename)
        print(f"{filename} has been deleted.")
    else:
        print(f"{filename} does not exist.")
    return f"Function executed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

def weekday_run():
    
    session = requests.Session()

    #Define headers
    headers = {
        'User-Agent': UserAgent().windows,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'en-US,en;q=0.9',
        'Connection': 'keep-alive',
        'DNT': '1',  # Do Not Track request header
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://amazon.in/',
    }

    session.cookies.set('session-id', '259-9131690-4965843', domain='.amazon.in')
    session.cookies.set('ubid-acbin', '262-3136477-5925466', domain='.amazon.in')
    session.cookies.set('x-acbin', 'oKxIUSyFo?JHcckDwUY?v@ZYnayTGSw8@bjN1GChjq0wUS@I8w31mOz?cHcIzdqI', domain='.amazon.in')

    sku_sheet = openpyxl.load_workbook("testSKU.xlsx").active
    marketplace_review_workbook = openpyxl.Workbook()
    default_sheet = marketplace_review_workbook.active
    marketplace_review_workbook.remove(default_sheet)
    marketplace_negative = marketplace_review_workbook.create_sheet(title="Negative review in last 3 days")
    head_ng = ["ECOM SKU","SKU Name","ASIN", "Marketplace", "Name", "Title","Review", "Rating", "Date", "Location", "Images"]

    marketplace_negative.append(head_ng)

    filename = f"Amazon_review_{current_time}.xlsx"

    for row in range(2,sku_sheet.max_row+1):
        sku = sku_sheet.cell(row=row,column=1).value
        amz_mark = sku_sheet.cell(row=row,column=3).value
        if amz_mark == 'Y':
            log_placeholder.write(f"Running SKU")
            asin = sku_sheet.cell(row=row,column=5).value
            print(f'/// ASIN {asin}')
            #Load asin page
            url = f'https://www.amazon.in/dp/{asin}'
            print(f'Load {url}...')
            response = session.get(url, headers=headers, verify=False)
            headers['Referer'] = url #save url as referrer for next query
            print(response)
            soup_mp = BeautifulSoup(response.content,"html.parser")
            sku_name = soup_mp.find('span', {'id': 'productTitle'}).get_text(strip=True)
            
            status_tag = soup_mp.find('span', {'class':'a-size-medium a-color-success'}) 
            if status_tag:
                status = status_tag.get_text(strip=True) 
            else:
                status = "Avaiable"
            print(sku_name)
            marketplace = "Amazon"
        
            # Find the div containing the rating histogram
            rating_histogram_div = soup_mp.find('div', class_='a-section a-spacing-none a-text-right aok-nowrap')

            # Initialize variables to store the ratings
            rating_5 = rating_4 = rating_3 = rating_2 = rating_1 = None

            if rating_histogram_div:
                # Extract the percentages from the span elements inside the div
                spans = rating_histogram_div.find_all('span', class_='_cr-ratings-histogram_style_histogram-column-space__RKUAd')
                
                # Store the ratings in corresponding variables
                if len(spans) >= 5:
                    rating_5 = spans[0].text.strip()  # First span: 51%
                    rating_4 = spans[1].text.strip()  # Second span: 18%
                    rating_3 = spans[2].text.strip()  # Third span: 3%
                    rating_2 = spans[3].text.strip()  # Fourth span: 2%
                    rating_1 = spans[4].text.strip()  # Fifth span: 26%
                
                # If there's additional text (like the final 51%) outside the spans, handle it
                last_rating = rating_histogram_div.contents[-1].strip()
                if last_rating and last_rating not in [rating_5, rating_4, rating_3, rating_2, rating_1]:
                    rating_5 = last_rating  # Assign it to rating_5 if not already set

            # Print the extracted ratings
            print(f'Rating 5: {rating_5}')
            print(f'Rating 4: {rating_4}')
            print(f'Rating 3: {rating_3}')
            print(f'Rating 2: {rating_2}')
            print(f'Rating 1: {rating_1}')
            avg_rating = soup_mp.find('span', {'data-hook': 'rating-out-of-text', 'aria-hidden': 'true'}).text.strip().split(' ')[0] if soup_mp.find('span', {'data-hook': 'rating-out-of-text', 'aria-hidden': 'true'}) else None
            print(avg_rating)
            number_of_rating = 0
            number_of_review = 0
            time.sleep(2)
            check_pr = soup_mp.find('span',{'data-hook' : 'total-review-count'})
            first_review_date = 101
            count_30 = 0
            count_365 = 0
            if check_pr:
                print("Running")
                page = 1
                while True:
                    url = f'https://www.amazon.in/product-reviews/{asin}/ref=cm_cr_arp_d_viewopt_srt?sortBy=recent&pageNumber={page}'
                    print(f'Page {page} | Load {url}...')
                    response_rw = session.get(url, headers=headers, verify=False)
                    headers['Referer'] = url  # save url as referrer for next query
                    time.sleep(2)
                    soup_rw = BeautifulSoup(response_rw.content,'html.parser')
                    reviews = soup_rw.findAll('li',{'class' : 'review aok-relative'})
                    # print(reviews)
                    if len(reviews) == 0:
                        break
                    
                    # no_review = soup_rw.find('div',{'class':'a-row a-spacing-base a-size-base'}).get_text(strip=True)

                    text = soup_rw.find('div', {'data-hook': 'cr-filter-info-review-rating-count'}).text.strip()
                    # Use regex to extract numbers
                    match = re.search(r'([\d,]+)\s+total rating(?:s)?,?\s+([\d,]+)?\s+with review(?:s)?', text)

                    number_of_rating = int(match.group(1).replace(',', ''))
                    number_of_review = int(match.group(2).replace(',', '')) if match.group(2) else 0
                    print(f"Number of ratings: {number_of_rating}")
                    print(f"Number of reviews: {number_of_review}")


                    print(number_of_rating)
                    print(number_of_review)
                    for review in reviews:

                        rating = int(float(review.find('span', class_='a-icon-alt').text.split(' ')[0]))
                        today = datetime.today()
                        review_text = review.find('span',{'data-hook':'review-body'}).find('span').text.strip() if review.find('span',{'data-hook':'review-body'}).find('span') else "NA"
                            review_date_text = review.find('span', {'data-hook': 'review-date'}).text.strip()

                        date_match = re.search(r'on (\d{1,2} \w+ \d{4})', review_date_text)
                        date = date_match.group(1) if date_match else None
                        if first_review_date == 101:
                            first_review_date = date
                        date_obj = datetime.strptime(date, "%d %B %Y")
                        if rating <= 3 and today - timedelta(days=3) <= date_obj:

                            name = review.find('span',{'class':'a-profile-name'}).text.strip()
                            title = review.find('a', {'data-hook': 'review-title'}).find_all('span')[-1].text.strip()
                            
                            # Use regex to extract location and date
                            location_match = re.search(r'Reviewed in (\w+)', review_date_text)
                            

                            # Assign to variables if matched
                            location = location_match.group(1) if location_match else None
                            
                            

                            # Calculate date ranges
                            thirty_days_ago = today - timedelta(days=30)
                            three_hundred_sixty_five_days_ago = today - timedelta(days=365)

                            # Check if the date falls within the last 30 days or 365 days
                            if thirty_days_ago <= date_obj <= today:
                                count_30 += 1
                            if three_hundred_sixty_five_days_ago <= date_obj <= today:
                                count_365 += 1
                            print(name+' '+str(rating)+' '+date+' '+location)
                            image_tags = review.find_all('img', class_='review-image-tile') if review.find_all('img', class_='review-image-tile') else None
                            all_links = "NA"
                            if image_tags is not None:
                                # Extract the 'src' links
                                image_links = [img['src'].replace('SY88', 'SL1600') for img in image_tags]

                                # Join the links into a single string with a delimiter (comma in this case)
                                all_links = ', '.join(image_links)

                            # Now you have all links in one variable
                            print(all_links)
                            marketplace_negative.append([
                                    sku,
                                    sku_name,
                                    asin,
                                    marketplace,
                                    name,
                                    title,
                                    review_text,
                                    rating,
                                    date,
                                    location,
                                    all_links
                                ])

                            next_page_check = soup_rw.find('li',{'class':'a-last'})
                            if next_page_check:
                                next_page = next_page_check.find('a')
                                if next_page:
                                    page +=1
                                else:
                                    break
                            else:
                                break
               
    
    marketplace_review_workbook.save(filename)

    with open(filename, mode='rb') as file:
        excel_base64 = base64.b64encode(file.read()).decode('utf-8')

    payload = {
        "personalizations": [
            {
                "recipient": "sparmar@godrej.com",
                "recipient_cc": ["parmarsachin707@gmail.com"]
            }
        ],
        "from": {
            "fromEmail": "info@godrejinterio.com",
        },
        "subject": "Amazon Ratings and Reviews",
        "content": "Hi, PFA the Ratings and Reviews Dump for Amazon",
        "templateId": 33201,
        "attachments": [
            {
                "fileContent": excel_base64,
                "fileName": "RatingsAndReviews.xlsx"
            }
        ]
    }

    headers = {
        "api_key": "50b39b265e9b8cca8a264652a27b57ef",
        "Content-Type": "application/json",
        "Accept": ""
    }
    url_email = "https://api.pepipost.com/v2/sendEmail"
    response_email = requests.post(url_email, json=payload, headers=headers)
    if os.path.exists(filename):
        os.remove(filename)
        print(f"{filename} has been deleted.")
    else:
        print(f"{filename} does not exist.")
    
    return f"Function executed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

# Streamlit app
st.title("Scheduled Function Execution")

# Display current status
status_placeholder = st.empty()
log_placeholder = st.empty()

while True:
    # Get current time in UTC
    current_time_utc = datetime.now(pytz.utc)

    # Convert to IST
    current_time_ist = current_time_utc.astimezone(pytz.timezone('Asia/Kolkata'))

    # Extract time component
    current_time = current_time_ist.time()
    start_time = datetime.strptime("00:06:00", "%H:%M:%S").time()
    end_time = datetime.strptime("00:25:00", "%H:%M:%S").time()
    is_monday = current_time_ist.weekday() == 0 
    if start_time <= current_time <= end_time:
        if is_monday:
            result = monday_run()
            log_placeholder.write(result)
            status_placeholder.info("Function executed successfully!")
            time.sleep(1000)  # Wait for 60 seconds
        else:
            result = weekday_run()
            log_placeholder.write(result)
            status_placeholder.info("Function executed successfully!")
            time.sleep(1000)  # Wait for 60 seconds
    else:
        status_placeholder.warning("Waiting for the time range...")
        time.sleep(1000)  # Check again in 10 seconds
